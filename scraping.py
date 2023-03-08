from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver  # Firefox/Gecko driver
from openpyxl import load_workbook  # write xlsx file
from datetime import datetime  # date management



# --- DEFINE GLOBAL
OK = 0
NEW = 'n'
UPDATED = 'u'
DELETED = 'd'
EMPTY = ""
equity_date_status_column_name = {'status': 'A', 'entry': 'B', 'out': 'C'}
equity_list_column_name = {'country': 'D', 'name': 'E', 'sector': 'F', 'capi': 'G',
						   '1st Jan': 'H', 'link': 'I', 'zb ref': 'J'}

# Directory profile and firefox options
profile_options = {
	"--profile": "/home/mouac/workspace/Bourse/scraping/Firefox_solenium_profile",
	"--headless": ""
}
# Dictionary of url to scrap
url2scrap = {
	"value Holland": "https://www.zonebourse.com/listes-investissement/valorisation/valorisation_pays_bas/",
	"value France":  "https://www.zonebourse.com/listes-investissement/valorisation/valorisation_france/",
	"value USA":     "https://www.zonebourse.com/listes-investissement/valorisation/valorisation_amerique/",
	"grow  Holland": "https://www.zonebourse.com/listes-investissement/croissance/croissance_pays_bas/",
	"grow  France":  "https://www.zonebourse.com/listes-investissement/croissance/croissance_france/",
	"grow  USA":     "https://www.zonebourse.com/listes-investissement/croissance/croissance_amerique/",
	"qual  France":  "https://www.zonebourse.com/listes-investissement/valeurs_de_qualite/valeurs_de_qualite_france/",
	"qual  Holland": "https://www.zonebourse.com/listes-investissement/valeurs_de_qualite/valeurs_de_qualite_paysbas/",
	"qual  USA":     "https://www.zonebourse.com/listes-investissement/valeurs_de_qualite/valeurs_de_qualite_amerique/",
	"mom   France":  "https://www.zonebourse.com/listes-investissement/valeurs_momentum/valeurs_momentum_france/",
	"mom   Holland": "https://www.zonebourse.com/listes-investissement/valeurs_momentum/valeurs_momentum_paysbas/",
	"mom   USA":     "https://www.zonebourse.com/listes-investissement/valeurs_momentum/valeurs_momentum_amerique/",
	"trend France":  "https://www.zonebourse.com/listes-investissement/valeurs_en_trend_following/trendfollowing_france/",
	"trend Holland": "https://www.zonebourse.com/listes-investissement/valeurs_en_trend_following/trendfollowing_pays_bas/",
	"trend USA":     "https://www.zonebourse.com/listes-investissement/valeurs_en_trend_following/trendfollowing_amerique/"
}

# Dictionary of sheet
scrap_in_sheet = {
	"euronext": ["value France", "value Holland",
				 "grow  France", "grow  Holland",
				 "qual  France", "qual  Holland",
				 "mom   France", "mom   Holland",
				 "trend France", "trend Holland"],
	"wall street": ["value USA", "grow  USA",
					"qual  USA", "mom   USA",
					"trend USA"]
}

"""
scrap_in_sheet = {
		"euronext": ["qual  france","qual  netherlands"],
	"wall street": []
}
"""
#  Select only country for wallstreet and euronext
selectable_country = ["us", "fr", "nl", "be"]
# --- END DEFINE ---

# Press Maj+F10 to execute it or replace it with your code.
# Presfrs Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# ---
# FUNCTION to initialize Firefox profile (FirefoxOptions, option_dictionary)
#   set the firefox web engine Options
#   dictionary key is the option
#   dictionary value is the option argument (if not needed value="")
#   RETURN status
def set_firefox_profile(fo, options):
	for key, opt in options.items():
		if key:
			fo.add_argument(key)
			if opt:
				fo.add_argument(opt)
	print(f'Firefox profile set with: {fo.arguments}')

# ---
# FUNCTION get_equity_style_invest (webdriver w, string si, equity_list[]):
#     This function gets the equity list of investment style in zonebourse website
#     ARGUMENTS:
#       webrdiver w: selenium webdrive instance
#       string si: investment style (use for score sheet)
#       [{}] equity_list: table of equity dict
#     RETURN table of dictionary
#       [{country, name, sector, capi, 1st January var, link, zb ref, invest style}]
def get_equity_style_invest(w, si, equity_list):
	try:
	# Get equity table and table length
		equity_table = w.find_element(By.ID, "ALNI0")  # id of the table to get
	except NoSuchElementException:
		w.quit()
		exit(-1)
	# line element of table
	equity_line = equity_table.find_elements(By.TAG_NAME, "tr")
	# Scan the table (start from line 2 because 1st line is head of column
	for i in range(2, len(equity_line)+1):
		j = i-2  # line number in equity_list
		equity_list_line = {}  # line to add in the equity list table
	# # # td[1] equity country
		equity_country_cell = w.find_element(By.XPATH, f'//*[@id="ALNI0"]/tbody/tr[{i}]/td[1]/img')
		equity_country_name = equity_country_cell.get_attribute("src")  # get the country img file name
		if equity_country_name[-6:-4] in selectable_country:
			equity_list_line['country'] = equity_country_name[-6:-4]
			# # # td[2] equity name
			equity_list_line['name'] = w.find_element(By.XPATH, f'//*[@id="ALNI0"]/tbody/tr[{i}]/td[2]').text
			# # # link to zb equity record and zonebourse equity code
			equity_zb_link_a = w.find_element(By.XPATH, f'//*[@id="ALNI0"]/tbody/tr[{i}]/td[2]/a')
			equity_list_line['link'] = equity_zb_link_a.get_attribute("href")
			equity_list_line['zb ref'] = equity_zb_link_a.get_attribute("codezb")
			# # # td[3] equity sector
			equity_list_line['sector'] = w.find_element(By.XPATH, f'//*[@id="ALNI0"]/tbody/tr[{i}]/td[3]').text
			# # # td[4] equity capital
			equity_list_line['capi'] = w.find_element(By.XPATH, f'//*[@id="ALNI0"]/tbody/tr[{i}]/td[4]').text
			# # # td[5] equity 1st january var
			equity_list_line['1st Jan'] = w.find_element(By.XPATH, f'//*[@id="ALNI0"]/tbody/tr[{i}]/td[5]').text

			# # # add investment style
			equity_list_line[si] = True
			### Add line in the equity table list without double entries
			if equity_list_line not in equity_list:
				equity_list.append(equity_list_line)
	return(equity_list)

# ---
# FUNCTION  read_zb_xls_file( worksheet ws):
#     This function read the zonebourse investment style list stored in one tab in the excel file
#     The function read only the zonebourse reference and return the list.
#     ARGUMENTS:
#       worksheet ws: the sheet to read
#     RETURN {} zonebourse reference (key) list with corresponding line number
def read_zb_ref_xls_file(ws):
	dico_zb_ref_file = {}
	for i in range(ws.max_row-1):  # -1 because start directly from 2nd line
		line = i+2  # sheet line start from 1 (+1) and we do not need the 1st linecolumn title (+1)
		# Store the value (i.e. zb ref) as key and the corresponding cell number (ex: "A1")
		zb_ref = ws[f'{equity_list_column_name["zb ref"]}{line}'].value
		dico_zb_ref_file[zb_ref] = line
	return(dico_zb_ref_file)

# ---
# FUNCTION write_zblist_xlsx ( [{}] nl, string sm, string fn):
#     This function write the zonebourse investment style list. One market place and one style by sheet
#     ARGUMENTS:
#       table of dictionary nl: list of equity get from zonebourse website
#       string sm: stock market place = tab name
#       string fn: file path name
#     RETURN OK
def write_zblist_xlsx(nl, sm, fn):
	wb = load_workbook(fn)
	sheet = wb[sm]
	#  date in/out  and status management
	now = datetime.now()
	current_date = now.strftime('%d.%m.%y')

	#  Read the excel file sheet to get zonebourse reference in table (with cell line)
	dico_zb_ref_file = read_zb_ref_xls_file(sheet)
	#  print(dico_zb_ref_file)

	#  for all the equity in the existing list (in xls file)
	for key in dico_zb_ref_file:
		zb_ref_found = False
		i = 0
		if len(nl) > 0:
			while i < len(nl):
				if nl[i]["zb ref"] == key:
					zb_ref_found = True
					break
				i += 1
			if zb_ref_found:
				# the equity is in the zb list that have just get from website
				# The equity is already in the xls file
				# Update the xls file, it is not "new" equity anymore
				sheet[f'{equity_date_status_column_name["status"]}{dico_zb_ref_file[nl[i]["zb ref"]]}'] = None
				#  print(f'Found - {nl[i]["name"]}')
				nl.pop(i)
			else:
				# The equity is already in the xls file and not in zb website
				# Update the xls file, it is not in the list anymore update out date
				if (sheet[f'{equity_date_status_column_name["out"]}{dico_zb_ref_file[key]}']).value == None:
					sheet[f'{equity_date_status_column_name["out"]}{dico_zb_ref_file[key]}'] = current_date
					sheet[f'{equity_date_status_column_name["status"]}{dico_zb_ref_file[key]}'] = DELETED

	# Now there are only new equity in the zb web list... to add in the excel file
	for i in range(0, len(nl)):
		j = sheet.max_row + 1  # exel line number
		#  Select only new equity from the zb list that have just got from website
		# add new equity in the xls file
		#  write status and date in excel sheet
		sheet[f'{equity_date_status_column_name["status"]}{j}'] = NEW
		sheet[f'{equity_date_status_column_name["entry"]}{j}'] = current_date
		sheet[f'{equity_date_status_column_name["out"]}{j}'] = None
		#  print(f'ADD - {nl[i]["name"]}')
		for key, value in equity_list_column_name.items():
			sheet[f'{equity_list_column_name[key]}{j}'] = nl[i][key]

	wb.save(fn)
	wb.close()
	return OK

# ---
# FUNCTION write_score_xlsx ( {{}} list, string sn, string fn):
#     This function write the list of all equity in dico and add for each one if in the invest style
#     ARGUMENTS:
#       dict of dict list: list of all equity with invest style indicator
#       string sn: score tab name
#       string fn: file path name
#     RETURN OK
def write_score_xlsx(list, sn, fn):
	# Open the sheet in xlsx file
	wb = load_workbook(fn)
	sheet = wb[sn]
	#  First delete all line in the score sheet
	while (sheet.max_row > 1):
		sheet.delete_rows(2)

	line = 2  # First line where to write the sheet
	# For all equity in the list
	for zb_ref in list:
		sheet[f'A{line}'] = list[zb_ref]['country']
		sheet[f'B{line}'].value = list[zb_ref]['name']
		sheet[f'B{line}'].hyperlink = list[zb_ref]['link']
		sheet[f'B{line}'].style = "Hyperlink"
		if 'value' in list[zb_ref]:
			sheet[f'C{line}'] = list[zb_ref]['value']
		if 'grow' in list[zb_ref]:
			sheet[f'D{line}'] = list[zb_ref]['grow']
		if 'qual' in list[zb_ref]:
			sheet[f'E{line}'] = list[zb_ref]['qual']
		if 'mom' in list[zb_ref]:
			sheet[f'F{line}'] = list[zb_ref]['mom']
		if 'trend' in list[zb_ref]:
			sheet[f'G{line}'] = list[zb_ref]['trend']
		sheet[f'H{line}'] = f'=SUM(C{line}:G{line})'
		line += 1
	wb.save(fn)
	wb.close()
	return OK

# ---
# FUNCTION merge_equity_list_score ( {[]} aels, [{}] el):
#     This function merge all sytle investment in one dictionnary with style information
#     ARGUMENTS:
#       dictionary of table: the key is zb reference and the table got all argmuents included all invest style
#       table of dictionary el: list of equity get from zonebourse website
#       string si: investment style
#     RETURN OK
def merge_equity_list_score(aels, el):
	#  first looking for equity list get from zb is in merged list
	for e in el:
		if e["zb ref"] not in aels:
			# if not in the dictionnary, add all data
			aels[e["zb ref"]] = e
		else:
			# if alredy in the dictionnary, update with the new invest style
			aels[e["zb ref"]].update(e)
	return aels


# ---
# FUNCTION zb_invest_style_scraping (webdriver w, str file_2w):
#     This function scrap all the zonebourse investment style equity list and write in xlsx file
#     one sheet by maketplace by style
#     One more sheet by marketplace to make filter by style
#     ARGUMENTS:
#       webrdiver w: selenium webdrive instance
#		string file_2w: xls file where to write result of the scraping
#     RETURN OK
def zb_invest_style_scraping(w, file_2w):
	#  make all tab sheet in the file
	for key in scrap_in_sheet:
		print("write marketplace ", key)
		all_equity_list_score = {}
		equity_list = []
		tab_sheet = scrap_in_sheet[key]
		type_invest = ""
		for ts in tab_sheet:
			ts_splitted = ts.split()
			if (ts_splitted[0] != type_invest) and (type_invest != ""):
				print("\twrite", sheet_name)
				write_zblist_xlsx(equity_list, sheet_name, file_2w)
				equity_list = []
			type_invest = ts_splitted[0]
			sheet_name = key + " " + ts_splitted[0]
			print("\t\tStart to scrap: ", url2scrap[ts])
			w.get(url2scrap[ts])
			equity_list = get_equity_style_invest(w, ts_splitted[0], equity_list)
			#  Merge list to do score sheet
			all_equity_list_score = merge_equity_list_score(all_equity_list_score, equity_list)
		print("\twrite", sheet_name)
		write_zblist_xlsx(equity_list, sheet_name, file_2w)
		#  Write score for all equity in the market place
		write_score_xlsx(all_equity_list_score, key, file_2w)

# ---
# FUNCTION main_scraping (str xls_file_2w):
#     main function to be call in your program
#     ARGUMENTS:
#       string : xls_file_2w xls file where to write results
#     RETURN OK
def main_scraping(xls_file_2w):
	# Init Firefox web browser profile to manage cookies and password
	firefox_options = FirefoxOptions()
	set_firefox_profile(firefox_options, profile_options)
	# webdriver creation
	driver = webdriver.Firefox(options=firefox_options)
	zb_invest_style_scraping(driver, xls_file_2w)

	driver.quit()
